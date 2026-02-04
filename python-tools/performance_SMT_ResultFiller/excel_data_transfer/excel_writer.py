"""
Excel写入模块 - 写入目标Excel文件
处理列映射、行查找和数据写入
"""
import re
import logging
from typing import List, Dict, Optional, Tuple, Union
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import (
    ANIMATION_SHEET_NAME,
    TCID_COLUMN,
    LEVEL_COLUMN,
    TEST_DEVICE_COLUMNS,
    COMPETITOR_DEVICE_COLUMNS,
    DEVICE_TYPE_TEST,
    DEVICE_TYPE_COMPETITOR,
    SLIDING_SUB_COLUMNS
)

logger = logging.getLogger(__name__)


def extract_folder_index(folder_name: str) -> Optional[int]:
    """
    从文件夹名称中提取索引数字
    例如: "1短信启动" -> 1, "2短信退出" -> 2

    Args:
        folder_name: 文件夹名称

    Returns:
        提取的索引数字，未找到则返回None
    """
    match = re.match(r'^(\d+)', folder_name)
    if match:
        return int(match.group(1))
    return None


def get_available_levels(sheet, column_mapping: Dict[str, int]) -> List[str]:
    """
    获取目标Excel中所有可用的用例等级

    Args:
        sheet: openpyxl工作表对象
        column_mapping: 列名到列索引的映射

    Returns:
        用例等级列表（去重后）
    """
    levels = set()

    if LEVEL_COLUMN not in column_mapping:
        logger.warning(f"未找到用例等级列: {LEVEL_COLUMN}")
        return []

    level_col_idx = column_mapping[LEVEL_COLUMN]

    for row in sheet.iter_rows(min_row=2, values_only=False):
        cell = row[level_col_idx - 1]
        if cell.value and str(cell.value).strip():
            levels.add(str(cell.value).strip())

    return sorted(list(levels))


def get_column_mapping(sheet) -> Dict[str, int]:
    """
    获取列名到列索引的映射
    对于滑动丢帧sheet：
    - 第1行包含：Tcid, 用例等级, Purpose, 测试机1, 测试机2, ...
    - 第2行包含：丢帧总数, 33ms次数, 50ms次数（在每个测试机列下面）

    Args:
        sheet: openpyxl工作表对象

    Returns:
        列名到列索引的字典
    """
    column_mapping = {}

    # 先检查是否是滑动丢帧sheet（通过检查第2行是否有"丢帧总数"等列名）
    header_row_1 = [cell.value for cell in sheet[1]]
    header_row_2 = [cell.value for cell in sheet[2]]

    # 如果第2行包含"丢帧总数"等列名，说明是滑动丢帧sheet
    if any(name in header_row_2 for name in ["丢帧总数", "33ms次数", "50ms次数"]):
        # 滑动丢帧sheet：需要读取两行的列名
        # 第1行：Tcid, 用例等级, Purpose, 测试机1, 测试机2, ...
        for i, cell in enumerate(sheet[1], start=1):
            if cell.value:
                column_mapping[cell.value] = cell.column
        # 第2行：丢帧总数, 33ms次数, 50ms次数
        for i, cell in enumerate(sheet[2], start=1):
            if cell.value and cell.value not in column_mapping:
                column_mapping[cell.value] = cell.column
    else:
        # 普通sheet：从第1行读取
        for cell in sheet[1]:
            if cell.value:
                column_mapping[cell.value] = cell.column

    return column_mapping


def find_row_by_tcid_index(sheet, tcid_col: str, index: int,
                           level_filter: Optional[str] = None) -> Optional[int]:
    """
    找到Tcid列下第index个有值的行号
    支持按用例等级筛选

    Args:
        sheet: openpyxl工作表对象
        tcid_col: Tcid列名或列索引
        index: 要查找的索引（从1开始）
        level_filter: 用例等级筛选条件，None表示不筛选

    Returns:
        行号，未找到则返回None
    """
    # 获取Tcid列的索引
    if isinstance(tcid_col, str):
        try:
            col_idx = column_index_from_string(tcid_col)
        except ValueError:
            # 尝试从列名映射中查找
            column_mapping = get_column_mapping(sheet)
            col_idx = column_mapping.get(tcid_col)
            if col_idx is None:
                logger.error(f"未找到列: {tcid_col}")
                return None
    else:
        col_idx = tcid_col

    # 获取用例等级列索引（如果需要筛选）
    level_col_idx = None
    if level_filter:
        column_mapping = get_column_mapping(sheet)
        if LEVEL_COLUMN in column_mapping:
            level_col_idx = column_mapping[LEVEL_COLUMN]
        else:
            logger.warning(f"用例等级筛选已指定，但未找到列: {LEVEL_COLUMN}")

    count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell_value = row[col_idx - 1].value

        # 检查是否符合用例等级筛选
        if level_filter and level_col_idx is not None:
            level_value = row[level_col_idx - 1].value
            if level_value != level_filter:
                continue

        if cell_value is not None:
            count += 1
            if count == index:
                return row_idx

    logger.warning(f"未找到Tcid列下第{index}个有值的行（筛选条件: 用例等级={level_filter}）")
    return None


def get_device_column_indices(column_mapping: Dict[str, int],
                              device_type: str) -> List[int]:
    """
    根据设备类型获取对应的列索引列表

    Args:
        column_mapping: 列名到列索引的映射
        device_type: 设备类型 ('test' 或 'competitor')

    Returns:
        列索引列表
    """
    if device_type == DEVICE_TYPE_TEST:
        columns = TEST_DEVICE_COLUMNS
    elif device_type == DEVICE_TYPE_COMPETITOR:
        columns = COMPETITOR_DEVICE_COLUMNS
    else:
        logger.error(f"未知的设备类型: {device_type}")
        return []

    indices = []
    for col_name in columns:
        if col_name in column_mapping:
            indices.append(column_mapping[col_name])
        else:
            logger.debug(f"列名 '{col_name}' 不在目标Excel中，跳过")

    return indices


def write_data_to_row(sheet, row_idx: int, data: List[Union[int, None]],
                      column_indices: List[int]) -> int:
    """
    将数据按顺序写入指定行的对应列
    None值会显示为字符串"None"

    Args:
        sheet: openpyxl工作表对象
        row_idx: 目标行号
        data: 要写入的数据列表（可能包含None）
        column_indices: 目标列索引列表

    Returns:
        实际写入的数据条数
    """
    written = 0
    for i, value in enumerate(data):
        if i >= len(column_indices):
            logger.warning(f"数据条数({len(data)})超过可用列数({len(column_indices)})，部分数据未写入")
            break

        col_idx = column_indices[i]
        # None值显示为字符串"None"
        write_value = "None" if value is None else value
        sheet.cell(row=row_idx, column=col_idx, value=write_value)
        written += 1

    return written


def save_workbook(wb, target_file) -> None:
    """
    保存Excel工作簿

    Args:
        wb: openpyxl工作簿对象
        target_file: 目标文件路径
    """
    try:
        wb.save(target_file)
        logger.info(f"数据已成功保存到 {target_file}")
    except Exception as e:
        logger.error(f"保存文件时出错: {e}")
        raise


def find_row_by_purpose(sheet, purpose_col: str, purpose_value: str,
                         level_filter: Optional[str] = None) -> Optional[int]:
    """
    根据Purpose列的值查找对应的行
    支持按用例等级筛选

    Args:
        sheet: openpyxl工作表对象
        purpose_col: Purpose列名
        purpose_value: 要匹配的Purpose值
        level_filter: 用例等级筛选条件

    Returns:
        行号，未找到则返回None
    """
    # 获取Purpose列的索引
    if isinstance(purpose_col, str):
        column_mapping = get_column_mapping(sheet)
        col_idx = column_mapping.get(purpose_col)
        if col_idx is None:
            logger.error(f"未找到列: {purpose_col}")
            return None
    else:
        col_idx = purpose_col

    # 获取用例等级列索引（如果需要筛选）
    level_col_idx = None
    if level_filter:
        column_mapping = get_column_mapping(sheet)
        if LEVEL_COLUMN in column_mapping:
            level_col_idx = column_mapping[LEVEL_COLUMN]
        else:
            logger.warning(f"用例等级筛选已指定，但未找到列: {LEVEL_COLUMN}")

    # 从第二行开始查找
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell_value = row[col_idx - 1].value

        # 检查是否符合用例等级筛选
        if level_filter and level_col_idx is not None:
            level_value = row[level_col_idx - 1].value
            if str(level_value).strip() != level_filter:
                continue

        # 匹配 Purpose 值（部分匹配，去除数字前缀后匹配）
        if cell_value and purpose_value in str(cell_value):
            return row_idx

    logger.warning(f"未找到 Purpose 值为 '{purpose_value}' 的行（筛选条件: 用例等级={level_filter}）")
    return None


def get_sliding_column_indices(column_mapping: Dict[str, int],
                               device_type: str) -> Dict[str, int]:
    """
    根据设备类型获取滑动丢帧对应的列索引
    返回 {main_N: 测试机N列索引, total_N: 丢帧总数列索引, count_33ms_N: 33ms次数列索引, count_50ms_N: 50ms次数列索引}

    Excel结构 (第1行是合并单元格，第2行是子表头):
    | 测试机1(合并3格) | 测试机2(合并3格) | ...
    | 丢帧总数 | 33ms次数 | 50ms次数 | 丢帧总数 | 33ms次数 | 50ms次数 | ...

    Args:
        column_mapping: 列名到列索引的映射（从第1行和第2行读取）
        device_type: 设备类型

    Returns:
        包含 main, total, count_33ms, count_50ms 的字典
    """
    if device_type == DEVICE_TYPE_TEST:
        device_columns = TEST_DEVICE_COLUMNS
    elif device_type == DEVICE_TYPE_COMPETITOR:
        device_columns = COMPETITOR_DEVICE_COLUMNS
    else:
        logger.error(f"未知的设备类型: {device_type}")
        return {}

    result = {}

    # 按顺序查找"测试机1", "测试机2", ... 的列索引（从第1行读取）
    for i, device_col_name in enumerate(device_columns):
        if device_col_name in column_mapping:
            main_col_idx = column_mapping[device_col_name]
            result[f'main_{i}'] = main_col_idx
            # 丢帧总数是测试机N合并单元格的第1个子列（与main列索引相同）
            result[f'total_{i}'] = main_col_idx
            # 33ms次数在右侧
            result[f'count_33ms_{i}'] = main_col_idx + 1
            # 50ms次数在33ms次数右侧
            result[f'count_50ms_{i}'] = main_col_idx + 2
        else:
            # 找不到该设备列，停止继续查找
            break

    logger.debug(f"滑动丢帧列索引映射: {result}")
    return result


def write_sliding_data_to_row(sheet, row_idx: int, data: dict,
                               column_indices: Dict[str, int]) -> int:
    """
    将滑动丢帧数据写入指定行
    - FrameOver33ms 第N个值 → 测试机N 列
    - 总丢帧数 第N个值 → 丢帧总数 列
    - FrameOver33ms >0 计数 → 测试机N 的 33ms次数 列
    - FrameOver50ms >0 计数 → 测试机N 的 50ms次数 列
    None值会显示为字符串"None"

    Excel结构: 测试机1 | 丢帧总数 | 33ms次数 | 50ms次数 | 测试机2 | ...

    Args:
        sheet: openpyxl工作表对象
        row_idx: 目标行号
        data: 包含 values_33ms, values_50ms, values_total, count_33ms, count_50ms 的字典
        column_indices: 列索引字典

    Returns:
        实际写入的数据条数
    """
    written = 0
    values_33ms = data.get('values_33ms', [])
    values_50ms = data.get('values_50ms', [])
    values_total = data.get('values_total', [])
    # 不再使用总count，而是为每个测试机计算各自的count

    # 遍历每个测试机位置，写入对应的数据
    i = 0
    while True:
        main_key = f'main_{i}'
        total_key = f'total_{i}'
        count_33ms_key = f'count_33ms_{i}'
        count_50ms_key = f'count_50ms_{i}'

        # 如果主列找不到，说明已处理完所有测试机
        if main_key not in column_indices:
            break

        main_col_idx = column_indices[main_key]

        # 写入 FrameOver33ms 第 i 个值到 测试机N 列
        if i < len(values_33ms):
            write_value = "None" if values_33ms[i] is None else values_33ms[i]
            sheet.cell(row=row_idx, column=main_col_idx, value=write_value)
            written += 1
        else:
            # 如果没有对应位置的数据，写入 "None"
            sheet.cell(row=row_idx, column=main_col_idx, value="None")
            written += 1

        # 写入 总丢帧数 第 i 个值到 丢帧总数 列
        if total_key in column_indices:
            if i < len(values_total):
                write_value = "None" if values_total[i] is None else values_total[i]
                sheet.cell(row=row_idx, column=column_indices[total_key], value=write_value)
                written += 1
            else:
                sheet.cell(row=row_idx, column=column_indices[total_key], value="None")
                written += 1

        # 写入 33ms次数：如果第i个值>0则写1，=0则写0，None则写"None"
        if count_33ms_key in column_indices:
            if i < len(values_33ms):
                if values_33ms[i] is None:
                    sheet.cell(row=row_idx, column=column_indices[count_33ms_key], value="None")
                elif values_33ms[i] > 0:
                    sheet.cell(row=row_idx, column=column_indices[count_33ms_key], value=1)
                else:
                    sheet.cell(row=row_idx, column=column_indices[count_33ms_key], value=0)
            else:
                sheet.cell(row=row_idx, column=column_indices[count_33ms_key], value="None")
            written += 1

        # 写入 50ms次数：如果第i个值>0则写1，=0则写0，None则写"None"
        if count_50ms_key in column_indices:
            if i < len(values_50ms):
                if values_50ms[i] is None:
                    sheet.cell(row=row_idx, column=column_indices[count_50ms_key], value="None")
                elif values_50ms[i] > 0:
                    sheet.cell(row=row_idx, column=column_indices[count_50ms_key], value=1)
                else:
                    sheet.cell(row=row_idx, column=column_indices[count_50ms_key], value=0)
            else:
                sheet.cell(row=row_idx, column=column_indices[count_50ms_key], value="None")
            written += 1

        i += 1

    return written
