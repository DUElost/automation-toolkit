"""
结果回填协调模块 - Top300 Result Filler
协调整个数据处理和回填流程
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .config import (
    LOAD_PREFIX,
    NO_LOAD_PREFIX,
    RESULT_SHEET_NAME,
)
from .excel_reader import get_valid_sheets, read_sheet_data, validate_file_exists
from .excel_writer import (
    create_new_sheet,
    fill_result_sheet,
    find_target_columns,
    save_workbook,
)
from .data_processor import generate_new_sheet_name, process_sheet_data

logger = logging.getLogger(__name__)


def process_source_file(file_path: Path, prefix: str) -> List[Dict]:
    """
    处理单个源文件（空载或负载）

    Args:
        file_path: 源文件路径
        prefix: 前缀（空载/负载）

    Returns:
        处理结果列表
    """
    results = []

    # 获取有效sheet列表
    valid_sheets = get_valid_sheets(file_path)
    if not valid_sheets:
        logger.warning(f"文件 {file_path.name} 中未找到有效sheet")
        return results

    # 处理每个sheet
    for idx, sheet_name in enumerate(valid_sheets, start=1):
        df = read_sheet_data(file_path, sheet_name)
        if df is None:
            continue

        # 处理数据
        process_result = process_sheet_data(df, sheet_name)

        if process_result["success"]:
            new_sheet_name = generate_new_sheet_name(sheet_name, prefix, idx)

            results.append({
                "original_sheet": sheet_name,
                "new_sheet_name": new_sheet_name,
                "zero_count": process_result["zero_count"],
                "average_value": process_result["column_average"],
                "processed_df": process_result["processed_df"],
            })

            logger.info(f"处理完成: {sheet_name} -> {new_sheet_name}, "
                       f"平均值={process_result['column_average']:.2f}")
        else:
            logger.error(f"处理sheet {sheet_name} 失败: {process_result.get('error')}")

    return results


def create_output_file(target_file: Path) -> Path:
    """
    创建输出文件（复制目标文件并添加时间戳）
    输出文件放置在main.py所在目录

    Args:
        target_file: 原目标文件路径

    Returns:
        新输出文件路径
    """
    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 获取main.py所在目录
    # result_filler.py 在 top300_processor/ 子目录下
    # 所以需要向上两级才能到达 main.py 所在目录
    current_dir = Path(__file__).resolve().parent
    main_dir = current_dir.parent  # 向上一级到 main.py 所在目录

    output_file = main_dir / f"{target_file.stem}_{timestamp}{target_file.suffix}"

    import shutil
    shutil.copy2(target_file, output_file)
    logger.info(f"创建输出文件: {output_file}")

    return output_file


def process_all_sources(
    no_load_file: Optional[Path],
    load_file: Optional[Path],
    target_file: Path,
    no_load_standard: str,
    load_standard: str,
) -> Dict:
    """
    主处理流程
    1. 读取空载文件（如果提供），获取有效sheet列表并处理
    2. 读取负载文件（如果提供），获取有效sheet列表并处理
    3. 复制目标文件为新文件（原文件名+时间戳）
    4. 在新文件中创建新sheet并写入数据
    5. 更新测试结果sheet

    Args:
        no_load_file: 空载文件路径（可选，None表示跳过）
        load_file: 负载文件路径（可选，None表示跳过）
        target_file: 目标Excel文件路径
        no_load_standard: 空载标准
        load_standard: 负载标准

    Returns:
        处理结果汇总字典
    """
    summary = {
        "success": False,
        "no_load_results": [],
        "load_results": [],
        "zero_counts": {},
        "averages": {},
        "output_file": None,
        "error": None
    }

    try:
        # 验证至少提供一个源文件
        if no_load_file is None and load_file is None:
            summary["error"] = "至少需要提供一个测试报告文件（空载或负载）"
            return summary

        # 验证目标文件存在
        if not validate_file_exists(target_file):
            summary["error"] = f"目标文件不存在: {target_file}"
            return summary

        # 验证空载文件存在（如果提供）
        if no_load_file is not None and not validate_file_exists(no_load_file):
            summary["error"] = f"空载文件不存在: {no_load_file}"
            return summary

        # 验证负载文件存在（如果提供）
        if load_file is not None and not validate_file_exists(load_file):
            summary["error"] = f"负载文件不存在: {load_file}"
            return summary

        # 创建输出文件（复制目标文件）
        output_file = create_output_file(target_file)
        summary["output_file"] = output_file

        # 初始化结果列表
        no_load_results = []
        load_results = []

        # 处理空载文件（如果提供）
        if no_load_file is not None:
            logger.info(f"开始处理空载文件: {no_load_file.name}")
            no_load_results = process_source_file(no_load_file, NO_LOAD_PREFIX)
            summary["no_load_results"] = no_load_results

            for result in no_load_results:
                summary["zero_counts"][result["new_sheet_name"]] = result["zero_count"]
                summary["averages"][result["new_sheet_name"]] = result["average_value"]
        else:
            logger.info("跳过空载文件处理")

        # 处理负载文件（如果提供）
        if load_file is not None:
            logger.info(f"开始处理负载文件: {load_file.name}")
            load_results = process_source_file(load_file, LOAD_PREFIX)
            summary["load_results"] = load_results

            for result in load_results:
                summary["zero_counts"][result["new_sheet_name"]] = result["zero_count"]
                summary["averages"][result["new_sheet_name"]] = result["average_value"]
        else:
            logger.info("跳过负载文件处理")

        # 打开输出文件（不是原目标文件）
        wb = load_workbook(output_file)

        # 创建新sheet并写入数据
        for result in no_load_results + load_results:
            create_new_sheet(
                wb,
                result["new_sheet_name"],
                result["processed_df"],
                result["zero_count"],
                result["average_value"]
            )

        # 更新测试结果sheet
        if RESULT_SHEET_NAME in wb.sheetnames:
            result_sheet = wb[RESULT_SHEET_NAME]
            column_map = find_target_columns(result_sheet)

            logger.info(f"探测到的列映射: {column_map}")

            required_columns = ["标准", "#1", "#2", "均值"]
            missing_columns = [col for col in required_columns if col not in column_map]

            if not missing_columns:
                logger.info("开始填写测试结果sheet...")
                fill_result_sheet(
                    result_sheet,
                    {
                        "no_load": no_load_results,
                        "load": load_results
                    },
                    {
                        "no_load": no_load_standard,
                        "load": load_standard
                    },
                    column_map
                )
            else:
                logger.warning(f"测试结果sheet缺少必要列，跳过回填。缺少列: {missing_columns}")
        else:
            logger.warning(f"未找到'{RESULT_SHEET_NAME}'sheet，跳过回填")

        # 保存输出文件
        save_workbook(wb, output_file)
        summary["success"] = True

        logger.info("数据处理完成")

    except Exception as e:
        summary["error"] = str(e)
        logger.error(f"处理过程中出错: {e}")

    return summary


def print_summary(summary: Dict) -> None:
    """
    打印处理结果摘要

    Args:
        summary: 处理结果汇总字典
    """
    print("\n" + "=" * 60)
    print("Top300 Result Filler - 处理结果")
    print("=" * 60)

    if summary["success"]:
        print("✓ 处理成功")

        # 空载结果
        no_load_results = summary.get("no_load_results", [])
        if no_load_results:
            print(f"\n空载文件处理 ({len(no_load_results)}个sheet):")
            for result in no_load_results:
                print(f"  - {result['new_sheet_name']}: "
                     f"零值个数={result['zero_count']}, "
                     f"平均值={result['average_value']:.2f}")
        else:
            print("\n空载文件: 跳过")

        # 负载结果
        load_results = summary.get("load_results", [])
        if load_results:
            print(f"\n负载文件处理 ({len(load_results)}个sheet):")
            for result in load_results:
                print(f"  - {result['new_sheet_name']}: "
                     f"零值个数={result['zero_count']}, "
                     f"平均值={result['average_value']:.2f}")
        else:
            print("\n负载文件: 跳过")

        # 输出文件信息
        if summary.get("output_file"):
            print(f"\n输出文件: {summary['output_file']}")

    else:
        print("✗ 处理失败")
        error = summary.get("error")
        if error:
            print(f"错误信息: {error}")

    print("=" * 60)
