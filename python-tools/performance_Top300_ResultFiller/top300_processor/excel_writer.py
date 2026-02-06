"""
Excel写入模块 - Top300 Result Filler
负责创建新sheet、写入数据、更新测试结果页
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import (
    COLUMN_AVERAGE,
    COLUMN_HASH_1,
    COLUMN_HASH_2,
    COLUMN_STANDARD,
    RESULT_SHEET_NAME,
    ROW_NO_LOAD_LABEL,
    ROW_LOAD_LABEL,
)

logger = logging.getLogger(__name__)


def save_workbook(wb, target_file: Path) -> None:
    """
    保存Excel工作簿

    Args:
        wb: openpyxl工作簿对象
        target_file: 目标文件路径
    """
    try:
        wb.save(target_file)
        logger.info(f"数据已成功保存到 {target_file}")
    except Exception as e:
        logger.error(f"保存文件时出错: {e}")
        raise


def find_target_columns(sheet) -> Dict[str, int]:
    """
    在目标sheet中探测关键列位置
    返回: {"标准": col_idx, "#1": col_idx, "#2": col_idx, "均值": col_idx}

    Args:
        sheet: openpyxl工作表对象

    Returns:
        列名到列索引的映射（1-based）
    """
    # 定义目标列名及其可能的变体
    target_columns = {
        COLUMN_STANDARD: ["标准", "Standard", "standard", "标准值"],
        COLUMN_HASH_1: ["#1", "1号", "No1", "第一台"],
        COLUMN_HASH_2: ["#2", "2号", "No2", "第二台"],
        COLUMN_AVERAGE: ["均值", "Average", "average", "平均值"],
    }

    column_map = {}

    # 扫描前10行来查找表头
    for header_row_idx in range(1, min(11, sheet.max_row + 1)):
        header_row = next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))

        # 记录原始行内容用于调试
        row_values = [str(v) if v is not None else "" for v in header_row]
        logger.debug(f"扫描第{header_row_idx}行: {row_values[:10]}")  # 只显示前10列

        for col_idx, value in enumerate(header_row, start=1):
            if value is None:
                continue

            # 去除空格后比较
            value_str = str(value).strip() if isinstance(value, str) else str(value)

            # 检查是否匹配任何目标列
            for target_key, possible_names in target_columns.items():
                if target_key in column_map:
                    continue  # 已找到该列
                if value_str in possible_names:
                    column_map[target_key] = col_idx
                    logger.info(f"找到列 '{target_key}' 在第{header_row_idx}行第{col_idx}列 (值: '{value_str}')")
                    break

        # 如果找到所有列就停止扫描
        if len(column_map) == len(target_columns):
            logger.info(f"在第{header_row_idx}行找到所有目标列")
            break

    logger.info(f"探测到目标列: {column_map}")
    return column_map


def find_rows_by_standard(sheet, column_map: Dict[str, int]) -> Dict[str, Optional[int]]:
    """
    根据"标准"列查找空载/负载行
    实际上"空载"/"负载"标签在A列（第1列），"标准"是B列的表头

    Args:
        sheet: openpyxl工作表对象
        column_map: 列映射字典

    Returns:
        {"no_load_row": row_idx, "load_row": row_idx}
    """
    result = {"no_load_row": None, "load_row": None}

    # "空载"/"负载"标签在第1列（A列）
    label_col = 1

    # 找到包含"标准"表头的行，确定数据区域起始行
    standard_col = column_map.get(COLUMN_STANDARD, 2)
    header_row_idx = None

    for row_idx in range(1, min(20, sheet.max_row + 1)):
        cell_value = sheet.cell(row=row_idx, column=standard_col).value
        if isinstance(cell_value, str) and cell_value.strip() == COLUMN_STANDARD:
            header_row_idx = row_idx
            logger.debug(f"找到'标准'列表头在第{row_idx}行第{standard_col}列")
            break

    # 从"标准"表头的下一行开始搜索
    if header_row_idx is None:
        # 如果找不到"标准"表头，尝试其他方式确定起始行
        # 直接扫描整个A列
        logger.warning("未找到'标准'列表头，扫描整行查找标签")
        start_row = 1
    else:
        start_row = header_row_idx + 1

    # 在A列中搜索"空载"/"负载"
    for row_idx in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=label_col).value

        if cell_value is None:
            continue

        if isinstance(cell_value, str):
            cell_value = cell_value.strip()

        if cell_value == ROW_NO_LOAD_LABEL:
            result["no_load_row"] = row_idx
            logger.debug(f"找到空载行: 第{row_idx}行 (A{row_idx})")
        elif cell_value == ROW_LOAD_LABEL:
            result["load_row"] = row_idx
            logger.debug(f"找到负载行: 第{row_idx}行 (A{row_idx})")

        # 如果都找到了就停止
        if result["no_load_row"] and result["load_row"]:
            break

    if result["no_load_row"] is None:
        logger.warning(f"未找到'{ROW_NO_LOAD_LABEL}'行")
    if result["load_row"] is None:
        logger.warning(f"未找到'{ROW_LOAD_LABEL}'行")

    logger.info(f"目标行查找结果: 空载行={result['no_load_row']}, 负载行={result['load_row']}")
    return result


def create_new_sheet(wb, sheet_name: str, df: pd.DataFrame, zero_count: int, column_average: float = None) -> float:
    """
    在目标工作簿中创建新sheet
    1. 写入所有原始数据（包括所有时间类型）
    2. "非首轮平均启动时间"列已由 data_processor 添加（仅符合条件的行有值）
    3. 在最后一行添加列平均值

    Args:
        wb: openpyxl工作簿对象
        sheet_name: 新sheet名称
        df: 处理后的DataFrame（包含所有原始数据）
        zero_count: AM且平均启动时间为0的个数
        column_average: 列平均值（如为None则从df中计算）

    Returns:
        列平均值（用于回填测试结果页）
    """
    try:
        # 删除已存在的sheet
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            logger.debug(f"删除已存在的sheet: {sheet_name}")

        # 创建新sheet
        ws = wb.create_sheet(title=sheet_name)

        # 写入数据（包括所有原始行）
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 获取非首轮平均启动时间列的位置
        non_first_col = len(df.columns)  # 最后一列

        # 如果没有提供平均值，从df中计算（仅有效值）
        if column_average is None:
            valid_values = df["非首轮平均启动时间"].dropna()
            valid_values = [v for v in valid_values if v != 0]
            column_average = sum(valid_values) / len(valid_values) if valid_values else 0.0

        # 在数据最后一行添加列平均值
        last_data_row = len(df) + 1  # +1因为有表头
        ws.cell(row=last_data_row + 1, column=non_first_col, value=column_average)

        total_rows = len(df)
        valid_rows = df["非首轮平均启动时间"].notna().sum()
        logger.info(f"创建sheet '{sheet_name}': 总数据={total_rows}行, "
                   f"计算数据={valid_rows}行, 零值个数={zero_count}, 列平均值={column_average:.2f}")

        return column_average

    except Exception as e:
        logger.error(f"创建sheet '{sheet_name}' 时出错: {e}")
        raise


def fill_result_sheet(sheet, results: Dict, standards: Dict[str, str], column_map: Dict[str, int]) -> None:
    """
    填写测试结果sheet

    Args:
        sheet: openpyxl工作表对象
        results: 处理结果字典
            {
                "no_load": [{"average_value": 123.45, "sheet_name": "空载1'0078"}, ...],
                "load": [{"average_value": 234.56, "sheet_name": "负载1'0078"}, ...]
            }
        standards: 标准值字典
            {"no_load": "用户输入的空载标准", "load": "用户输入的负载标准"}
        column_map: 列映射字典
    """
    logger.info(f"开始填写测试结果sheet, column_map={column_map}")

    rows = find_rows_by_standard(sheet, column_map)

    # 填写标准值
    if rows["no_load_row"]:
        sheet.cell(row=rows["no_load_row"], column=column_map[COLUMN_STANDARD], value=standards["no_load"])
        logger.info(f"填写空载标准: '{standards['no_load']}' 到第{rows['no_load_row']}行")
    if rows["load_row"]:
        sheet.cell(row=rows["load_row"], column=column_map[COLUMN_STANDARD], value=standards["load"])
        logger.info(f"填写负载标准: '{standards['load']}' 到第{rows['load_row']}行")

    # 填写空载数据
    no_load_results = results.get("no_load", [])
    logger.info(f"空载结果数量: {len(no_load_results)}")
    if rows["no_load_row"] and len(no_load_results) >= 2:
        row = rows["no_load_row"]
        val1 = no_load_results[0].get("average_value", 0)
        val2 = no_load_results[1].get("average_value", 0)

        sheet.cell(row=row, column=column_map[COLUMN_HASH_1], value=val1)
        sheet.cell(row=row, column=column_map[COLUMN_HASH_2], value=val2)
        sheet.cell(row=row, column=column_map[COLUMN_AVERAGE], value=(val1 + val2) / 2)

        logger.info(f"填写空载行(第{row}行): #1={val1:.2f}, #2={val2:.2f}, 均值={(val1 + val2) / 2:.2f}")
    else:
        if not rows["no_load_row"]:
            logger.warning("跳过空载数据填写: 未找到空载行")
        if len(no_load_results) < 2:
            logger.warning(f"跳过空载数据填写: 空载结果数量不足({len(no_load_results)} < 2)")

    # 填写负载数据
    load_results = results.get("load", [])
    logger.info(f"负载结果数量: {len(load_results)}")
    if rows["load_row"] and len(load_results) >= 2:
        row = rows["load_row"]
        val1 = load_results[0].get("average_value", 0)
        val2 = load_results[1].get("average_value", 0)

        sheet.cell(row=row, column=column_map[COLUMN_HASH_1], value=val1)
        sheet.cell(row=row, column=column_map[COLUMN_HASH_2], value=val2)
        sheet.cell(row=row, column=column_map[COLUMN_AVERAGE], value=(val1 + val2) / 2)

        logger.info(f"填写负载行(第{row}行): #1={val1:.2f}, #2={val2:.2f}, 均值={(val1 + val2) / 2:.2f}")
    else:
        if not rows["load_row"]:
            logger.warning("跳过负载数据填写: 未找到负载行")
        if len(load_results) < 2:
            logger.warning(f"跳过负载数据填写: 负载结果数量不足({len(load_results)} < 2)")
