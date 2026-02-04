"""
核心转移逻辑模块
协调整个数据转移流程
"""
import shutil
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from .config import (
    SOURCE_DIR,
    SOURCE_DIR_SLIDING,
    SOURCE_DIR_COMPETITOR,
    SOURCE_DIR_SLIDING_COMPETITOR,
    TARGET_FILE,
    BACKUP_DIR,
    TEMPLATE_FILE,
    ANIMATION_SHEET_NAME,
    TCID_COLUMN,
    LEVEL_COLUMN,
    SLIDING_SHEET_NAME,
    SLIDING_MATCH_COLUMN,
    SLIDING_SUB_COLUMNS,
    DEVICE_TYPE_TEST,
    DEVICE_TYPE_COMPETITOR,
    DEVICE_TYPE_LABELS,
    DATA_TYPE_ANIMATION,
    DATA_TYPE_SLIDING,
    DATA_TYPE_LABELS
)
from .excel_reader import collect_all_drop_frames, collect_sliding_drop_frames
from .excel_writer import (
    extract_folder_index,
    get_column_mapping,
    find_row_by_tcid_index,
    find_row_by_purpose,
    get_device_column_indices,
    get_sliding_column_indices,
    write_data_to_row,
    write_sliding_data_to_row,
    save_workbook,
    get_available_levels
)

logger = logging.getLogger(__name__)


def create_backup(file_path: Path) -> Path:
    """
    创建文件备份

    Args:
        file_path: 原文件路径

    Returns:
        备份文件路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = file_path.parent / f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    logger.info(f"已创建备份: {backup_path}")
    return backup_path


def copy_template_to_current(template_file: Path, output_file: Path = None) -> Path:
    """
    从backup文件夹复制模板文件到当前目录

    Args:
        template_file: 模板文件路径
        output_file: 输出文件路径，如果为None则在当前目录生成

    Returns:
        生成的文件路径
    """
    if not template_file.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    if output_file is None:
        # 在当前目录生成，文件名添加时间戳
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = Path.cwd() / f"{template_file.stem}_{timestamp}{template_file.suffix}"

    shutil.copy2(template_file, output_file)
    logger.info(f"已从模板复制文件: {template_file} -> {output_file}")
    return output_file


def get_sorted_subfolders(source_dir: Path) -> list:
    """
    获取源目录下的所有子文件夹，并按名称中的数字排序

    Args:
        source_dir: 源目录路径

    Returns:
        排序后的子文件夹列表
    """
    if not source_dir.exists():
        logger.error(f"源目录不存在: {source_dir}")
        return []

    subfolders = [d for d in source_dir.iterdir() if d.is_dir()]

    # 按文件夹名称中的数字排序
    def sort_key(folder):
        idx = extract_folder_index(folder.name)
        return idx if idx is not None else 999

    return sorted(subfolders, key=sort_key)


def process_single_folder(folder: Path, sheet, column_mapping: dict,
                          device_type: str, level_filter: str = None) -> dict:
    """
    处理单个文件夹的数据转移

    Args:
        folder: 源文件夹路径
        sheet: openpyxl工作表对象
        column_mapping: 列名到列索引的映射
        device_type: 设备类型
        level_filter: 用例等级筛选条件

    Returns:
        处理结果字典
    """
    result = {
        'folder_name': folder.name,
        'success': False,
        'row_index': None,
        'data_count': 0,
        'written_count': 0,
        'error': None
    }

    try:
        # 提取文件夹索引
        folder_index = extract_folder_index(folder.name)
        if folder_index is None:
            result['error'] = f"无法从文件夹名称 '{folder.name}' 中提取索引"
            logger.warning(result['error'])
            return result

        # 找到目标行（带用例等级筛选）
        row_idx = find_row_by_tcid_index(sheet, TCID_COLUMN, folder_index, level_filter)
        if row_idx is None:
            result['error'] = f"未找到Tcid列下第{folder_index}个有值的行（筛选: {level_filter or '无'}）"
            logger.warning(result['error'])
            return result

        result['row_index'] = row_idx

        # 收集该文件夹下所有数据
        drop_frames = collect_all_drop_frames(folder)
        if not drop_frames:
            result['error'] = f"文件夹 '{folder.name}' 中无有效数据"
            logger.info(result['error'])
            return result

        result['data_count'] = len(drop_frames)

        # 获取设备列索引
        column_indices = get_device_column_indices(column_mapping, device_type)
        if not column_indices:
            result['error'] = f"未找到设备类型 '{device_type}' 对应的列"
            logger.warning(result['error'])
            return result

        # 写入数据
        written = write_data_to_row(sheet, row_idx, drop_frames, column_indices)
        result['written_count'] = written
        result['success'] = True

        logger.info(f"文件夹 '{folder.name}' 处理完成: 写入 {written} 条数据到第 {row_idx} 行")

    except Exception as e:
        result['error'] = str(e)
        logger.error(f"处理文件夹 '{folder.name}' 时出错: {e}")

    return result


def process_folder_transfer(source_dir: Path = SOURCE_DIR,
                            target_file: Path = TARGET_FILE,
                            device_type: str = DEVICE_TYPE_TEST,
                            level_filter: str = None,
                            create_backup_flag: bool = True) -> dict:
    """
    主转移逻辑

    Args:
        source_dir: 源目录路径
        target_file: 目标Excel文件路径
        device_type: 设备类型 ('test' 或 'competitor')
        level_filter: 用例等级筛选条件 (None表示不筛选)
        create_backup_flag: 是否创建备份

    Returns:
        处理结果汇总字典
    """
    summary = {
        'success': False,
        'total_folders': 0,
        'processed_folders': 0,
        'total_data_written': 0,
        'failed_folders': [],
        'backup_file': None,
        'level_filter': level_filter,
        'error': None
    }

    try:
        # 检查源目录
        if not source_dir.exists():
            summary['error'] = f"源目录不存在: {source_dir}"
            logger.error(summary['error'])
            return summary

        # 检查目标文件
        if not target_file.exists():
            summary['error'] = f"目标文件不存在: {target_file}"
            logger.error(summary['error'])
            return summary

        # 创建备份
        if create_backup_flag:
            backup_path = create_backup(target_file)
            summary['backup_file'] = str(backup_path)

        # 加载目标文件
        wb = load_workbook(target_file)
        if ANIMATION_SHEET_NAME not in wb.sheetnames:
            summary['error'] = f"工作表 '{ANIMATION_SHEET_NAME}' 不存在"
            logger.error(summary['error'])
            return summary

        sheet = wb[ANIMATION_SHEET_NAME]

        # 获取列映射
        column_mapping = get_column_mapping(sheet)
        logger.info(f"检测到 {len(column_mapping)} 个列")

        # 获取子文件夹列表
        subfolders = get_sorted_subfolders(source_dir)
        summary['total_folders'] = len(subfolders)

        logger.info(f"找到 {len(subfolders)} 个子文件夹")
        logger.info(f"设备类型: {DEVICE_TYPE_LABELS.get(device_type, device_type)}")
        if level_filter:
            logger.info(f"用例等级筛选: {level_filter}")

        # 处理每个文件夹
        for folder in subfolders:
            result = process_single_folder(folder, sheet, column_mapping, device_type, level_filter)

            if result['success']:
                summary['processed_folders'] += 1
                summary['total_data_written'] += result['written_count']
            else:
                summary['failed_folders'].append({
                    'folder': folder.name,
                    'error': result.get('error')
                })

        # 保存文件
        save_workbook(wb, target_file)
        summary['success'] = True

    except Exception as e:
        summary['error'] = str(e)
        logger.error(f"处理过程中出错: {e}")

    return summary


def print_summary(summary: dict) -> None:
    """
    打印处理结果摘要

    Args:
        summary: 处理结果字典
    """
    print("\n" + "=" * 50)
    print("数据转移完成")
    print("=" * 50)

    if summary['success']:
        print(f"[OK] 处理成功")
        print(f"  总文件夹数: {summary['total_folders']}")
        print(f"  成功处理: {summary['processed_folders']}")
        print(f"  总写入数据: {summary['total_data_written']} 条")
        if summary.get('level_filter'):
            print(f"  用例等级筛选: {summary['level_filter']}")

        if summary['failed_folders']:
            print(f"\n跳过的文件夹 ({len(summary['failed_folders'])}):")
            for item in summary['failed_folders']:
                print(f"  - {item['folder']}: {item['error']}")

        if summary['backup_file']:
            print(f"\n备份文件: {summary['backup_file']}")
    else:
        print(f"[FAIL] 处理失败: {summary['error']}")

    print("=" * 50)


def get_available_levels_from_file(target_file: Path) -> list:
    """
    从目标Excel文件获取可用的用例等级列表

    Args:
        target_file: 目标Excel文件路径

    Returns:
        可用的用例等级列表
    """
    try:
        wb = load_workbook(target_file)
        if ANIMATION_SHEET_NAME not in wb.sheetnames:
            return []

        sheet = wb[ANIMATION_SHEET_NAME]
        column_mapping = get_column_mapping(sheet)
        return get_available_levels(sheet, column_mapping)
    except Exception as e:
        logger.error(f"获取用例等级列表时出错: {e}")
        return []


def process_sliding_single_folder(folder: Path, sheet, column_mapping: dict,
                                   device_type: str, level_filter: str = None) -> dict:
    """
    处理单个滑动丢帧文件夹的数据转移

    Args:
        folder: 源文件夹路径
        sheet: openpyxl工作表对象
        column_mapping: 列名到列索引的映射
        device_type: 设备类型
        level_filter: 用例等级筛选条件

    Returns:
        处理结果字典
    """
    result = {
        'folder_name': folder.name,
        'success': False,
        'row_index': None,
        'values_33ms': [],
        'values_50ms': [],
        'values_total': [],
        'count_33ms': 0,
        'count_50ms': 0,
        'error': None
    }

    try:
        # 提取文件夹名称（去掉数字前缀，如 "3.WhatsApp个人对话列表滑动" -> "WhatsApp个人对话列表滑动"）
        folder_name = folder.name
        if '.' in folder_name:
            folder_name = folder_name.split('.', 1)[1] if folder_name.split('.', 1)[1] else folder_name

        # 根据 Purpose 列查找目标行
        row_idx = find_row_by_purpose(sheet, SLIDING_MATCH_COLUMN, folder_name, level_filter)
        if row_idx is None:
            result['error'] = f"未找到 Purpose 列值为 '{folder_name}' 的行（筛选: {level_filter or '无'}）"
            logger.warning(result['error'])
            return result

        result['row_index'] = row_idx

        # 收集该文件夹下的滑动丢帧数据
        sliding_data = collect_sliding_drop_frames(folder)

        # 即使数据为None也继续处理，写入"None"字符串
        if sliding_data.get("found_file") is None:
            result['error'] = f"文件夹 '{folder.name}' 中未找到Excel文件"
            logger.info(result['error'])
            return result

        # 记录值（即使为None）
        result['values_33ms'] = sliding_data.get('values_33ms', [])
        result['values_50ms'] = sliding_data.get('values_50ms', [])
        result['values_total'] = sliding_data.get('values_total', [])
        result['count_33ms'] = sliding_data.get('count_33ms', 0)
        result['count_50ms'] = sliding_data.get('count_50ms', 0)

        values_33ms = result['values_33ms']
        if not values_33ms or all(v is None for v in values_33ms):
            logger.info(f"文件夹 '{folder.name}' 中 FrameOver33ms 列为空，将写入 'None'")
        else:
            logger.info(f"文件夹 '{folder.name}' 中 FrameOver33ms 数据: {len(values_33ms)} 个值, 计数: {result['count_33ms']}")

        # 获取设备列索引
        column_indices = get_sliding_column_indices(column_mapping, device_type)
        if not column_indices:
            result['error'] = f"未找到设备类型 '{device_type}' 对应的列"
            logger.warning(result['error'])
            return result

        # 写入数据
        written = write_sliding_data_to_row(sheet, row_idx, sliding_data, column_indices)
        result['written'] = written
        result['success'] = True

        logger.info(f"文件夹 '{folder.name}' 处理完成: 写入数据到第 {row_idx} 行")

    except Exception as e:
        result['error'] = str(e)
        logger.error(f"处理文件夹 '{folder.name}' 时出错: {e}")

    return result


def process_sliding_transfer(source_dir: Path = SOURCE_DIR,
                              target_file: Path = TARGET_FILE,
                              device_type: str = DEVICE_TYPE_TEST,
                              level_filter: str = None,
                              create_backup_flag: bool = True) -> dict:
    """
    滑动丢帧主转移逻辑

    Args:
        source_dir: 源目录路径
        target_file: 目标Excel文件路径
        device_type: 设备类型 ('test' 或 'competitor')
        level_filter: 用例等级筛选条件 (None表示不筛选)
        create_backup_flag: 是否创建备份

    Returns:
        处理结果汇总字典
    """
    summary = {
        'success': False,
        'total_folders': 0,
        'processed_folders': 0,
        'total_values_33ms': 0,
        'total_values_total': 0,
        'total_counts_33ms': 0,
        'total_counts_50ms': 0,
        'failed_folders': [],
        'backup_file': None,
        'level_filter': level_filter,
        'error': None
    }

    try:
        # 检查源目录
        if not source_dir.exists():
            summary['error'] = f"源目录不存在: {source_dir}"
            logger.error(summary['error'])
            return summary

        # 检查目标文件
        if not target_file.exists():
            summary['error'] = f"目标文件不存在: {target_file}"
            logger.error(summary['error'])
            return summary

        # 创建备份
        if create_backup_flag:
            backup_path = create_backup(target_file)
            summary['backup_file'] = str(backup_path)

        # 加载目标文件
        wb = load_workbook(target_file)
        if SLIDING_SHEET_NAME not in wb.sheetnames:
            summary['error'] = f"工作表 '{SLIDING_SHEET_NAME}' 不存在"
            logger.error(summary['error'])
            return summary

        sheet = wb[SLIDING_SHEET_NAME]

        # 获取列映射
        column_mapping = get_column_mapping(sheet)
        logger.info(f"检测到 {len(column_mapping)} 个列")

        # 获取子文件夹列表
        subfolders = [d for d in source_dir.iterdir() if d.is_dir()]
        summary['total_folders'] = len(subfolders)

        logger.info(f"找到 {len(subfolders)} 个子文件夹")
        logger.info(f"设备类型: {DEVICE_TYPE_LABELS.get(device_type, device_type)}")
        if level_filter:
            logger.info(f"用例等级筛选: {level_filter}")

        # 处理每个文件夹
        for folder in subfolders:
            result = process_sliding_single_folder(folder, sheet, column_mapping, device_type, level_filter)

            if result['success']:
                summary['processed_folders'] += 1
                summary['total_values_33ms'] += len(result.get('values_33ms', []))
                summary['total_values_total'] += len(result.get('values_total', []))
                summary['total_counts_33ms'] += result.get('count_33ms', 0)
                summary['total_counts_50ms'] += result.get('count_50ms', 0)
            else:
                summary['failed_folders'].append({
                    'folder': folder.name,
                    'error': result.get('error')
                })

        # 保存文件
        save_workbook(wb, target_file)
        summary['success'] = True

    except Exception as e:
        summary['error'] = str(e)
        logger.error(f"处理过程中出错: {e}")

    return summary


def print_sliding_summary(summary: dict) -> None:
    """
    打印滑动丢帧处理结果摘要

    Args:
        summary: 处理结果字典
    """
    print("\n" + "=" * 50)
    print("滑动丢帧数据转移完成")
    print("=" * 50)

    if summary['success']:
        print(f"[OK] 处理成功")
        print(f"  总文件夹数: {summary['total_folders']}")
        print(f"  成功处理: {summary['processed_folders']}")
        print(f"  总写入33ms数据: {summary['total_values_33ms']} 个")
        print(f"  总写入丢帧总数: {summary['total_values_total']} 个")
        print(f"  总写入33ms计数: {summary['total_counts_33ms']} 次")
        print(f"  总写入50ms计数: {summary['total_counts_50ms']} 次")
        if summary.get('level_filter'):
            print(f"  用例等级筛选: {summary['level_filter']}")

        if summary['failed_folders']:
            print(f"\n跳过的文件夹 ({len(summary['failed_folders'])}):")
            for item in summary['failed_folders']:
                print(f"  - {item['folder']}: {item['error']}")

        if summary['backup_file']:
            print(f"\n备份文件: {summary['backup_file']}")
    else:
        print(f"[FAIL] 处理失败: {summary['error']}")

    print("=" * 50)


def batch_process(device_types: list = None, data_types: list = None,
                  animation_source: Path = None, sliding_source: Path = None,
                  animation_source_comp: Path = None, sliding_source_comp: Path = None,
                  target_file: Path = None, level_filter: str = None,
                  use_template: bool = True) -> dict:
    """
    批量处理函数：支持同时处理多种设备类型和数据类型

    Args:
        device_types: 设备类型列表，如 ['test'] 或 ['test', 'competitor']
        data_types: 数据类型列表，如 ['animation'] 或 ['animation', 'sliding']
        animation_source: 测试机动效丢帧源目录
        sliding_source: 测试机滑动丢帧源目录
        animation_source_comp: 竞品机动效丢帧源目录
        sliding_source_comp: 竞品机滑动丢帧源目录
        target_file: 目标Excel文件路径，None表示从模板生成新文件
        level_filter: 用例等级筛选条件
        use_template: 是否使用模板生成新文件

    Returns:
        批量处理结果汇总字典
    """
    # 默认值处理
    if device_types is None or len(device_types) == 0:
        device_types = [DEVICE_TYPE_TEST]
    if data_types is None or len(data_types) == 0:
        data_types = [DATA_TYPE_ANIMATION]
    if animation_source is None:
        animation_source = SOURCE_DIR
    if sliding_source is None:
        sliding_source = SOURCE_DIR_SLIDING
    if animation_source_comp is None:
        animation_source_comp = SOURCE_DIR_COMPETITOR
    if sliding_source_comp is None:
        sliding_source_comp = SOURCE_DIR_SLIDING_COMPETITOR

    batch_summary = {
        'success': False,
        'use_template': use_template,
        'target_file': str(target_file) if target_file else None,
        'results': [],
        'total_processed': 0,
        'total_failed': 0,
        'error': None
    }

    try:
        # 确定目标文件
        if use_template:
            if not TEMPLATE_FILE.exists():
                batch_summary['error'] = f"模板文件不存在: {TEMPLATE_FILE}"
                logger.error(batch_summary['error'])
                return batch_summary

            # 从模板生成新文件
            if target_file is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                target_file = Path.cwd() / f"{TEMPLATE_FILE.stem}_{timestamp}{TEMPLATE_FILE.suffix}"
            else:
                target_file = Path(target_file)

            target_file = copy_template_to_current(TEMPLATE_FILE, target_file)
            batch_summary['target_file'] = str(target_file)
            logger.info(f"已从模板生成文件: {target_file}")
        else:
            if target_file is None:
                target_file = TARGET_FILE
            else:
                target_file = Path(target_file)

            if not target_file.exists():
                batch_summary['error'] = f"目标文件不存在: {target_file}"
                logger.error(batch_summary['error'])
                return batch_summary

        # 批量处理所有组合
        for data_type in data_types:
            for device_type in device_types:
                # 如果参数为空，跳过该部分处理
                if device_type is None or data_type is None:
                    logger.info(f"跳过设备类型={device_type}, 数据类型={data_type}（参数为空）")
                    continue

                # 根据设备类型选择对应的源目录
                if device_type == DEVICE_TYPE_COMPETITOR:
                    current_animation_source = animation_source_comp
                    current_sliding_source = sliding_source_comp
                else:
                    current_animation_source = animation_source
                    current_sliding_source = sliding_source

                result = {
                    'data_type': data_type,
                    'data_type_label': DATA_TYPE_LABELS.get(data_type, data_type),
                    'device_type': device_type,
                    'device_type_label': DEVICE_TYPE_LABELS.get(device_type, device_type),
                    'success': False,
                    'error': None,
                    'source_dir': str(current_animation_source) if data_type == DATA_TYPE_ANIMATION else str(current_sliding_source)
                }

                try:
                    if data_type == DATA_TYPE_SLIDING:
                        summary = process_sliding_transfer(
                            source_dir=current_sliding_source,
                            target_file=target_file,
                            device_type=device_type,
                            level_filter=level_filter,
                            create_backup_flag=False  # 批量处理时不创建多个备份
                        )
                        result['summary'] = summary
                        result['success'] = summary['success']
                        if summary['success']:
                            batch_summary['total_processed'] += 1
                        else:
                            batch_summary['total_failed'] += 1
                            result['error'] = summary.get('error')
                    else:  # DATA_TYPE_ANIMATION
                        summary = process_folder_transfer(
                            source_dir=current_animation_source,
                            target_file=target_file,
                            device_type=device_type,
                            level_filter=level_filter,
                            create_backup_flag=False
                        )
                        result['summary'] = summary
                        result['success'] = summary['success']
                        if summary['success']:
                            batch_summary['total_processed'] += 1
                        else:
                            batch_summary['total_failed'] += 1
                            result['error'] = summary.get('error')

                except Exception as e:
                    result['error'] = str(e)
                    batch_summary['total_failed'] += 1
                    logger.error(f"处理 {data_type}/{device_type} 时出错: {e}")

                batch_summary['results'].append(result)

        batch_summary['success'] = batch_summary['total_failed'] == 0

    except Exception as e:
        batch_summary['error'] = str(e)
        logger.error(f"批量处理过程中出错: {e}")

    return batch_summary


def print_batch_summary(summary: dict) -> None:
    """
    打印批量处理结果摘要

    Args:
        summary: 批量处理结果字典
    """
    print("\n" + "=" * 60)
    print("批量数据转移完成")
    print("=" * 60)

    if summary.get('use_template'):
        print(f"[模板] 已从模板生成新文件")
    print(f"目标文件: {summary.get('target_file', 'N/A')}")

    if summary['success']:
        print(f"[OK] 全部处理成功")
        print(f"  成功处理: {summary['total_processed']} 项")
        print(f"  失败: {summary['total_failed']} 项")
    else:
        print(f"[FAIL] 部分或全部处理失败")

    print("\n详细结果:")
    for result in summary.get('results', []):
        status = "[OK]" if result['success'] else "[FAIL]"
        print(f"  {status} {result['data_type_label']} - {result['device_type_label']}")
        if result['success'] and 'summary' in result:
            s = result['summary']
            if s.get('processed_folders'):
                print(f"       处理文件夹: {s['processed_folders']}/{s['total_folders']}")
        elif not result['success']:
            print(f"       错误: {result.get('error', 'Unknown')}")

    print("=" * 60)
